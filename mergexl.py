import sys
import os
import openpyxl

xl_dir = sys.argv[1]

xl_filenames = []
for filename in os.listdir(xl_dir):
    if filename.endswith('.xlsx'):
        xl_filenames.append(filename)

xl_filenames.sort()

main_wb = openpyxl.load_workbook(f'{xl_dir}/{xl_filenames[0]}')
main_ws = main_wb.active

for filename in xl_filenames[1:]:
    wb = openpyxl.load_workbook(f'{xl_dir}/{filename}')
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        main_ws.append(row)

main_wb.save(filename=f'{xl_dir}/{xl_filenames[0]}-merged.xlsx')