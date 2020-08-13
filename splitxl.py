import sys
import os
import openpyxl

def chunks(seq, size):
    for i in range(0, len(seq), size):
        yield seq[i: i + size]

xl_path, count = sys.argv[1:]
filename = os.path.basename(xl_path)
file_basename, ext = os.path.splitext(filename)
splitted_files_path = f'{os.path.splitext(xl_path)[0]}-splitted'
if not os.path.exists(splitted_files_path):
    os.mkdir(splitted_files_path)


wb = openpyxl.load_workbook(xl_path)
ws = wb.active
total_count = ws.max_row
lines_per_file = total_count // int(count)
print(f'Total lines: {total_count}, split count: {count}, {lines_per_file} lines per file')
header = [row for row in ws.iter_rows(min_row=1, max_row=1, values_only=True)][0]
print(f'header: {header}')
chunked_rows = list(chunks([row for row in ws.iter_rows(min_row=2, values_only=True)], lines_per_file))
if len(chunked_rows) > int(count):
    last_chunk = chunked_rows.pop() + chunked_rows.pop()
    chunked_rows.append(last_chunk)

for i, rows in enumerate(chunked_rows):
    new_wb = openpyxl.Workbook()
    active_ws = new_wb.active
    active_ws.append(header)
    for row in rows:
        active_ws.append(row)
    new_wb.save(filename=f'{splitted_files_path}/{file_basename}-{i+1}{ext}')